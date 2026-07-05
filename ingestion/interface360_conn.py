"""Interface 360 connector — parses the 24-column interfaces.xlsx.

Uses ProjectResolver for source/target project mapping. Pure parsing; the
loader persists. Y/N normalization and feed-type canonicalization applied.
"""
from __future__ import annotations
import hashlib
import os
import re
import logging
from typing import Optional

from .base import BaseConnector
from .project_resolver import ProjectResolver
from .interface360_model import Interface, Interface360Bundle

log = logging.getLogger("cp.interface360")

# header text (row 1) -> field. Matched by normalized substring.
_CANON_FEED_TYPES = ["Batch", "REST API", "SOAP", "SFTP", "MQ", "Kafka",
                     "DB Link", "File", "GraphQL"]


def _yn(v) -> str:
    s = str(v).strip().lower() if v is not None else ""
    return "Y" if s in ("y", "yes", "true", "1") else "N"


def _canon_feed_type(v) -> Optional[str]:
    if not v:
        return None
    s = str(v).strip().lower()
    for t in _CANON_FEED_TYPES:
        if t.lower() in s:
            return t
    return str(v).strip()


def _split_routing(routing: Optional[str]) -> list[str]:
    if not routing:
        return []
    parts = re.split(r"-+>", routing)
    return [p.strip() for p in parts if p.strip()]


class Interface360Connector(BaseConnector):
    name = "interface360"

    def __init__(self, xlsx_path: str, resolver: ProjectResolver):
        self.xlsx_path = xlsx_path
        self.resolver = resolver

    @classmethod
    def from_env(cls) -> "Interface360Connector":
        return cls(
            xlsx_path=os.environ["INTERFACE360_XLSX_PATH"],
            resolver=ProjectResolver.from_env(),
        )

    def parse(self) -> Interface360Bundle:
        from openpyxl import load_workbook
        wb = load_workbook(self.xlsx_path, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        return self._parse_rows(rows)

    def _parse_rows(self, rows: list) -> Interface360Bundle:
        if not rows:
            return Interface360Bundle()
        # positional mapping (24 columns, fixed order per spec)
        bundle = Interface360Bundle()
        for i, r in enumerate(rows[1:], start=1):
            if not any(c is not None and str(c).strip() for c in r):
                continue
            g = lambda idx: r[idx] if idx < len(r) and r[idx] is not None else None
            src = (str(g(8)) if g(8) else "") or ""
            tgt = (str(g(10)) if g(10) else "") or ""
            iface = Interface(
                interface_id=hashlib.sha1(
                    f"{g(4)}|{g(5)}|{src}|{tgt}|{i}".encode()).hexdigest()[:16],
                domain=_s(g(0)), date_of_update=_s(g(1)), scope=_s(g(2)),
                update_owner=_s(g(3)), application=_s(g(4)),
                integration_name=_s(g(5)), description=_s(g(6)),
                feed_type=_canon_feed_type(g(7)),
                source_system=_s(g(8)), source_party=_s(g(9)),
                target_system=_s(g(10)), target_party=_s(g(11)),
                direction=_s(g(12)), direct_feed=_yn(g(13)),
                feed_routing=_s(g(14)), intraday=_yn(g(15)),
                eod_overnight=_yn(g(16)), frequency=_s(g(17)),
                extract_type=_s(g(18)), app_owner=_s(g(19)),
                migration_flag=_yn(g(20)), type_app_extract=_s(g(21)),
                improvement=_s(g(22)), notes=_s(g(23)),
                source_project_id=self.resolver.resolve_for_interface_system(src),
                target_project_id=self.resolver.resolve_for_interface_system(tgt),
                routing_hops=_split_routing(_s(g(14))),
            )
            bundle.interfaces.append(iface)
        log.info("interface360: parsed %d interfaces", len(bundle.interfaces))
        return bundle


def _s(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None
