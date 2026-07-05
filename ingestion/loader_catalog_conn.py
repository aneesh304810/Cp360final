"""
Loader catalog connector — reads the loader Excel (templates arranged per your
spec: loader template + error template + schema definition, plus which feeds are
inbound vs outbound per loader).

Excel structure:
  - Sheet 1 (index): one row per loader —
      Loader Name | Template | Domain | Error Template | Inbound Feeds | Outbound Feeds
  - Optional per-loader sheet (name == loader name): schema def / validation rows.

Writes:
  - loader_catalog
  - pipeline_members rows are NOT written here (the BA composes pipelines); but the
    inbound/outbound feed columns are captured so the builder can pre-suggest them.
"""
from __future__ import annotations
import logging
import os
from pathlib import Path

from openpyxl import load_workbook

log = logging.getLogger("cp.loader_catalog")


def _norm(s):
    return str(s).strip() if s is not None else ""


def _header_map(ws):
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        return {(_norm(h).lower()): i for i, h in enumerate(row) if h is not None}
    return {}


class LoaderCatalogConnector:
    def __init__(self, xlsx_path: str, resolver, platform_id="SWP"):
        self.xlsx_path = xlsx_path
        self.resolver = resolver
        self.platform_id = platform_id

    @classmethod
    def from_env(cls, resolver):
        path = os.environ.get("LOADER_CATALOG_XLSX")
        return cls(path, resolver) if path else None

    def parse(self) -> dict:
        wb = load_workbook(self.xlsx_path, data_only=True, read_only=True)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        hm = _header_map(ws)

        loaders = []
        suggested = []   # (loader_id, direction, feed_id) hints for the builder
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            def c(*names):
                for n in names:
                    if n in hm and hm[n] < len(row):
                        return _norm(row[hm[n]])
                return ""
            lname = c("loader name", "loader", "name")
            if not lname:
                continue
            loader_id = lname.replace(" ", "_")[:200]
            # schema def: from a per-loader sheet if present, else the cell
            schema_def = c("schema def", "schema definition", "schema", "layout")
            if lname in sheets:
                lws = wb[lname]
                rows = []
                for lrow in lws.iter_rows(min_row=1, values_only=True):
                    if lrow and any(lrow):
                        rows.append(" | ".join(_norm(x) for x in lrow if x is not None))
                if rows:
                    schema_def = "\n".join(rows)[:4000]
            loaders.append({
                "loader_id": loader_id, "loader_name": lname[:400],
                "template": c("template", "loader template")[:200] or None,
                "schema_def": (schema_def or "")[:4000] or None,
                "error_template": c("error template", "error handling", "error")[:400] or None,
                "validation_rules": c("validation", "validation rules", "rules")[:4000] or None,
                "business_domain": c("domain")[:120] or None,
                "project_id": self.resolver.resolve_for_swp_feed(lname),
                "source_xlsx": self.xlsx_path,
            })
            # parse inbound/outbound feed lists (comma/semicolon separated)
            for direction, hdrs in (("inbound", ("inbound feeds", "inbound", "in feeds")),
                                     ("outbound", ("outbound feeds", "outbound", "out feeds"))):
                raw = c(*hdrs)
                for feed in _split(raw):
                    suggested.append({"loader_id": loader_id, "direction": direction,
                                      "feed_id": feed.replace(" ", "_")[:200]})
        wb.close()
        log.info("loader_catalog: %d loaders, %d feed-suggestions", len(loaders), len(suggested))
        return {"loaders": loaders, "suggested": suggested}

    def load(self, loader, bundle):
        for l in bundle["loaders"]:
            loader._merge("loader_catalog", ("loader_id",), l)
        loader.commit()


def _split(raw):
    if not raw:
        return []
    for sep in (";", ","):
        if sep in raw:
            return [x.strip() for x in raw.split(sep) if x.strip()]
    return [raw.strip()] if raw.strip() else []
