"""PII Classification connector — parses PII_Attributes_List.xlsx.

One attribute groups many components. Builds the normalized PII dictionary used
by the matcher.
"""
from __future__ import annotations
import os
import logging
from dataclasses import dataclass, field
from typing import Optional

from .base import BaseConnector

log = logging.getLogger("cp.pii")

SENSITIVITY_ORDINAL = {
    "Client Level Sensitive": 4,
    "Client and Employee Sensitive": 4,
    "Account Level Sensitive": 3,
    "Asset Level Sensitive": 2,
}


def normalize(name: str) -> str:
    return (name or "").lower().replace("_", "").replace(" ", "").replace("-", "")


@dataclass
class PiiClassification:
    pii_component_normalized: str
    pii_component: str
    pii_attribute: str
    sensitivity_category: str
    sensitivity_level: int
    source_xlsx_path: Optional[str] = None


@dataclass
class PiiBundle:
    classifications: list[PiiClassification] = field(default_factory=list)


class PiiClassificationConnector(BaseConnector):
    name = "pii_classification"

    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path

    @classmethod
    def from_env(cls) -> "PiiClassificationConnector":
        return cls(xlsx_path=os.environ["PII_ATTRIBUTES_PATH"])

    def parse(self) -> PiiBundle:
        from openpyxl import load_workbook
        wb = load_workbook(self.xlsx_path, data_only=True, read_only=True)
        ws = wb["PII Attributes"] if "PII Attributes" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        return self._parse_rows(rows)

    def _parse_rows(self, rows: list) -> PiiBundle:
        bundle = PiiBundle()
        seen = set()
        for r in rows[1:] if rows else []:
            if not r or r[0] is None:
                continue
            attribute = _s(r[0])
            component = _s(r[1]) if len(r) > 1 else None
            category = _s(r[2]) if len(r) > 2 else None
            if not component or not category:
                continue
            norm = normalize(component)
            if norm in seen:
                continue
            seen.add(norm)
            bundle.classifications.append(PiiClassification(
                pii_component_normalized=norm,
                pii_component=component, pii_attribute=attribute,
                sensitivity_category=category,
                sensitivity_level=SENSITIVITY_ORDINAL.get(category, 1),
                source_xlsx_path=self.xlsx_path,
            ))
        log.info("pii: parsed %d classifications", len(bundle.classifications))
        return bundle


def _s(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None
